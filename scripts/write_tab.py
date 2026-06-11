#!/usr/bin/env python3
"""
Write a CSV into a named tab in an Excel workbook.

Usage:
  python3 write_tab.py --file <path.xlsx> --tab <sheet-name> --csv <path.csv>

Creates the file if it does not exist. Replaces the tab if it already exists.
Preserves all other tabs in the workbook.
"""

import argparse
import csv
import sys
import json
from pathlib import Path


def write_tab(xlsx_path: str, tab_name: str, csv_path: str) -> dict:
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip3 install openpyxl"}

    path = Path(xlsx_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(str(path))
    else:
        wb = Workbook()
        default = wb.active
        if default and default.title in ("Sheet", "Sheet1"):
            wb.remove(default)

    if tab_name in wb.sheetnames:
        del wb[tab_name]

    ws = wb.create_sheet(title=tab_name)

    with open(csv_path, encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))

    if not rows:
        wb.save(str(path))
        return {"file": str(path), "tab": tab_name, "rows": 0, "columns": 0}

    header = rows[0]
    for col_idx, value in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")

    for row_idx, row in enumerate(rows[1:], 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    wb.save(str(path))

    return {
        "file": str(path),
        "tab": tab_name,
        "rows": len(rows) - 1,
        "columns": len(header),
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to .xlsx file (created if missing)")
    parser.add_argument("--tab", required=True, help="Tab (sheet) name to write")
    parser.add_argument("--csv", required=True, help="Path to input CSV file")
    args = parser.parse_args()

    result = write_tab(args.file, args.tab, args.csv)
    print(json.dumps(result, indent=2))
    if "error" in result:
        sys.exit(1)
